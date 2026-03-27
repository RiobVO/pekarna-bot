import asyncio
from aiogram import Router, F, BaseMiddleware
from aiogram.types import Message, CallbackQuery
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from sqlalchemy import select
from database.db import session_maker
from database.models import Product, Order, OrderItem, Client, ClientPrice
from keyboards import manage_menu, order_status_keyboard, admin_reply_menu
from config import is_admin
from datetime import datetime, timedelta
import config

# Все кнопки reply-меню админа
ADMIN_MENU_BUTTONS = {
    "📦 Новые заказы", "⚙️ Активные заказы", "📋 История заказов",
    "🍞 Управление меню", "🔍 Найти заказ", "📅 Заказы по дате",
    "📊 Статистика", "📥 Экспорт в Excel", "📢 Рассылка",
    "💰 Договорные цены", "👥 Клиенты", "⏸ Пауза", "▶️ Возобновить",
}

# {chat_id: [bot_message_id, ...]} — предыдущие ответы бота
_last_bot_msgs: dict[int, list[int]] = {}
# {user_id} — кто сейчас обрабатывается (защита от параллельных нажатий)
_processing: set[int] = set()


def _track_msg(chat_id: int, msg: Message):
    """Запоминает ID сообщения бота чтобы удалить при следующем нажатии"""
    _last_bot_msgs.setdefault(chat_id, []).append(msg.message_id)


class AdminMenuMiddleware(BaseMiddleware):
    """
    Перехватывает нажатия кнопок reply-меню админа и делает три вещи:
    1. Удаляет сообщение пользователя (кнопку) из чата
    2. Игнорирует нажатие если предыдущее ещё обрабатывается
    3. Сбрасывает FSM-состояние и удаляет предыдущие ответы бота
    """
    async def __call__(self, handler, event: Message, data: dict):
        if not isinstance(event, Message):
            return await handler(event, data)
        if not is_admin(event.from_user.id) or event.text not in ADMIN_MENU_BUTTONS:
            return await handler(event, data)

        user_id = event.from_user.id

        # Удаляем сообщение пользователя (кнопку) сразу
        try:
            await event.delete()
        except Exception:
            pass

        # Уже обрабатывается — игнорируем
        if user_id in _processing:
            return

        # Сбрасываем FSM (иначе кнопка меню попадёт в FSM-хендлер)
        state: FSMContext = data.get("state")
        if state:
            await state.clear()

        # Удаляем предыдущие ответы бота
        bot = data.get("bot")
        chat_id = event.chat.id
        for msg_id in _last_bot_msgs.pop(chat_id, []):
            try:
                await bot.delete_message(chat_id, msg_id)
            except Exception:
                pass

        _processing.add(user_id)
        try:
            return await handler(event, data)
        finally:
            _processing.discard(user_id)



router = Router()
router.message.middleware(AdminMenuMiddleware())  # inner: state доступен в data


def _requisites_text(client) -> str:
    parts = []
    if client.inn:
        parts.append(f"🏦 ИНН: {client.inn}")
    if client.bank_account:
        mfo_str = f" | МФО: {client.mfo}" if client.mfo else ""
        parts.append(f"💳 Р/С: {client.bank_account}{mfo_str}")
    if client.bank_name:
        parts.append(f"🏛 {client.bank_name}")
    return "\n".join(parts) + "\n" if parts else ""


class AdminState(StatesGroup):
    adding_product_name = State()   # ввод названия продукта
    adding_product_price = State()  # ввод цены продукта
    deleting_product = State()      # ввод номера для удаления
    rejecting_order = State()

@router.message(F.text == "/admin")
async def admin_start(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("❌ У вас нет доступа.")
        return
    await message.answer(
        "👋 Привет, диспетчер!",
        reply_markup=admin_reply_menu()
    )


# ─── УПРАВЛЕНИЕ ЗАКАЗАМИ ───────────────────────────────


@router.callback_query(F.data == "new_orders")
async def new_orders(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    async with session_maker() as session:
        orders = await session.execute(
            select(Order).where(Order.status == "new")
        )
        orders = orders.scalars().all()

        if not orders:
            await callback.message.answer("📭 Новых заказов нет.")
            return

        for order in orders:
            client = await session.get(Client, order.client_id)
            items = await session.execute(
                select(OrderItem).where(OrderItem.order_id == order.id)
            )
            items = items.scalars().all()

            type_text = "🚚 Доставка" if order.delivery_type == "delivery" else "🏪 Самовывоз"
            address_line = f"\n📍 {order.delivery_address}" if order.delivery_type == "delivery" and order.delivery_address else ""
            requisites = _requisites_text(client)
            text = (
                f"📦 ЗАКАЗ #{order.id}\n"
                f"👤 {client.contact_name}\n"
                f"🏢 {client.organization}\n"
                f"📞 {client.phone}\n"
                f"{requisites}"
                f"━━━━━━━━━━━━━━━\n"
            )
            for item in items:
                text += f"• {item.product_name} — {item.quantity} шт\n"
            text += (
                f"━━━━━━━━━━━━━━━\n"
                f"{type_text}{address_line}\n"
                f"📅 {order.delivery_date} в {order.delivery_time}"
            )
            from keyboards import order_action
            await callback.message.answer(text, reply_markup=order_action(order.id))


@router.callback_query(F.data.startswith("accept_"))
async def accept_order(callback: CallbackQuery, bot):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    order_id = int(callback.data.split("_")[1])

    async with session_maker() as session:
        order = await session.get(Order, order_id)
        order.status = "accepted"
        await session.commit()
        client = await session.get(Client, order.client_id)
        telegram_id = client.telegram_id

    sent = await bot.send_message(
        telegram_id,
        f"✅ Заказ #{order_id} принят! Начинаем готовить 👨‍🍳"
    )
    async with session_maker() as session:
        order = await session.get(Order, order_id)
        order.status_message_id = sent.message_id
        await session.commit()

    try:
        await callback.message.delete()
    except Exception:
        pass

@router.callback_query(F.data.startswith("reject_"))
async def reject_order(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    order_id = int(callback.data.split("_")[1])
    await state.update_data(rejecting_order_id=order_id, rejecting_message_id=callback.message.message_id)
    await state.set_state(AdminState.rejecting_order)
    await callback.message.answer(
        f"❌ Причина отклонения заказа #{order_id}:\n"
        f"(или напишите «-» чтобы не указывать)"
    )


@router.message(AdminState.rejecting_order)
async def reject_with_reason(message: Message, state: FSMContext, bot):
    data = await state.get_data()
    order_id = data["rejecting_order_id"]
    reason = message.text.strip()
    await state.clear()

    async with session_maker() as session:
        order = await session.get(Order, order_id)
        order.status = "rejected"
        await session.commit()
        client = await session.get(Client, order.client_id)
        telegram_id = client.telegram_id

    if reason == "-":
        client_text = f"❌ Ваш заказ #{order_id} отклонён. Свяжитесь с нами."
    else:
        client_text = f"❌ Ваш заказ #{order_id} отклонён.\n📝 Причина: {reason}"

    await bot.send_message(telegram_id, client_text)
    await message.answer(f"✅ Заказ #{order_id} отклонён.")

    rejecting_message_id = data.get("rejecting_message_id")
    if rejecting_message_id:
        try:
            await bot.delete_message(message.chat.id, rejecting_message_id)
        except Exception:
            pass


@router.callback_query(F.data == "all_orders")
async def all_orders(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    async with session_maker() as session:
        orders = await session.execute(select(Order))
        orders = orders.scalars().all()

        if not orders:
            await callback.message.answer("📭 Заказов пока нет.")
            return

        status_map = {
            "new":       "🟡 Новый",
            "accepted":  "✅ Принят",
            "rejected":  "❌ Отклонён",
            "cancelled": "🚫 Отменён",
            "preparing": "👨‍🍳 Готовится",
            "on_way":    "🚚 В пути",
            "ready":     "📦 Готов к выдаче",
            "delivered": "✅ Доставлен",
            "given":     "✅ Выдан",
        }

        chunks = ["📋 ВСЕ ЗАКАЗЫ:\n\n"]
        for order in orders:
            client = await session.get(Client, order.client_id)
            type_icon = "🚚" if order.delivery_type == "delivery" else "🏪"
            line = (
                f"{status_map.get(order.status)} #{order.id} — "
                f"{client.contact_name} — "
                f"{type_icon} {order.delivery_date}\n"
            )
            if len(chunks[-1]) + len(line) > 3500:
                chunks.append("")
            chunks[-1] += line

    for chunk in chunks:
        if chunk.strip():
            await callback.message.answer(chunk)



@router.callback_query(F.data == "active_orders")
async def active_orders(callback: CallbackQuery):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    active = ["accepted", "preparing", "on_way", "ready"]

    async with session_maker() as session:
        orders = await session.execute(
            select(Order).where(Order.status.in_(active))
        )
        orders = orders.scalars().all()

        if not orders:
            await callback.message.answer("📭 Активных заказов нет.")
            return

        for order in orders:
            client = await session.get(Client, order.client_id)
            items = await session.execute(
                select(OrderItem).where(OrderItem.order_id == order.id)
            )
            items = items.scalars().all()

            type_text = "🚚 Доставка" if order.delivery_type == "delivery" else "🏪 Самовывоз"
            address_line = f"\n📍 {order.delivery_address}" if order.delivery_type == "delivery" and order.delivery_address else ""
            status_map = {
                "accepted":  "✅ Принят",
                "preparing": "👨‍🍳 Готовится",
                "on_way":    "🚚 В пути",
                "ready":     "📦 Готов к выдаче",
            }
            requisites = _requisites_text(client)

            text = (
                f"📦 Заказ #{order.id} — {status_map.get(order.status)}\n"
                f"👤 {client.contact_name}\n"
                f"🏢 {client.organization}\n"
                f"{requisites}"
                f"━━━━━━━━━━━━━━━\n"
            )
            for item in items:
                text += f"• {item.product_name} — {item.quantity} шт\n"
            text += (
                f"━━━━━━━━━━━━━━━\n"
                f"{type_text}{address_line}\n"
                f"📅 {order.delivery_date} в {order.delivery_time}"
            )

            await callback.message.answer(
                text,
                reply_markup=order_status_keyboard(order.id, order.delivery_type, order.status)
            )



# ─── УПРАВЛЕНИЕ МЕНЮ ───────────────────────────────────


@router.callback_query(F.data == "manage_menu")
async def manage_menu_handler(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    try:
        await callback.message.edit_text(
            "🍞 Управление меню:",
            reply_markup=manage_menu()
        )
    except Exception:
        pass
    await callback.answer()


@router.callback_query(F.data == "list_products")
async def list_products(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return

    async with session_maker() as session:
        products = await session.execute(select(Product))
        products = products.scalars().all()

    if not products:
        await callback.answer("😔 Меню пустое.", show_alert=True)
        return

    text = "📋 СПИСОК ПРОДУКТОВ:\n\n"
    for i, p in enumerate(products, 1):
        text += f"{i}. {p.name} — {p.price} сум\n"

    try:
        await callback.message.edit_text(
            text,
            reply_markup=manage_menu()
        )
    except Exception:
        pass
    await callback.answer()


@router.callback_query(F.data == "add_product")
async def add_product(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return
    await state.set_state(AdminState.adding_product_name)
    _track_msg(callback.message.chat.id, await callback.message.answer("Введите название продукта:"))


@router.message(AdminState.adding_product_name)
async def enter_product_name(message: Message, state: FSMContext):
    await state.update_data(product_name=message.text)
    await state.set_state(AdminState.adding_product_price)
    _track_msg(message.chat.id, await message.answer("Введите цену (в сумах):"))


@router.message(AdminState.adding_product_price)
async def enter_product_price(message: Message, state: FSMContext):
    try:
        price = int(message.text)
        if price <= 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите правильную цену числом")
        return

    data = await state.get_data()
    product_name = data["product_name"]

    async with session_maker() as session:
        existing = await session.execute(select(Product).where(Product.name == product_name))
        if existing.scalar_one_or_none():
            await message.answer(f"❌ Продукт '{product_name}' уже существует. Введите другое название:")
            await state.set_state(AdminState.adding_product_name)
            return
        session.add(Product(name=product_name, price=price))
        await session.commit()

    await state.clear()
    _track_msg(message.chat.id, await message.answer(f"✅ Продукт '{product_name}' добавлен!"))
    _track_msg(message.chat.id, await message.answer("🍞 Управление меню:", reply_markup=manage_menu()))


@router.callback_query(F.data == "delete_product")
async def delete_product(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    async with session_maker() as session:
        products = await session.execute(select(Product))
        products = products.scalars().all()

    if not products:
        await callback.message.answer("😔 Меню пустое.")
        return

    text = "Введите номер продукта для удаления:\n\n"
    for i, p in enumerate(products, 1):
        text += f"{i}. {p.name} — {p.price} сум\n"

    await state.update_data(products=[(p.id, p.name) for p in products])
    await state.set_state(AdminState.deleting_product)
    _track_msg(callback.message.chat.id, await callback.message.answer(text))



@router.message(AdminState.deleting_product)
async def confirm_delete(message: Message, state: FSMContext):
    data = await state.get_data()
    products = data.get("products")
    if not products:
        await state.clear()
        return

    try:
        index = int(message.text) - 1
        if index < 0 or index >= len(products):
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите правильный номер")
        return

    product_id, product_name = products[index]

    async with session_maker() as session:
        product = await session.get(Product, product_id)
        await session.delete(product)
        await session.commit()

    await state.clear()
    _track_msg(message.chat.id, await message.answer(f"✅ Продукт '{product_name}' удалён!"))
    _track_msg(message.chat.id, await message.answer("🍞 Управление меню:", reply_markup=manage_menu()))


@router.callback_query(F.data == "admin_back")
async def admin_back(callback: CallbackQuery):
    try:
        await callback.message.edit_text(
            "🍞 Управление меню:",
            reply_markup=manage_menu()
        )
    except Exception:
        pass
    await callback.answer()

# ─── СТАТУСЫ ЗАКАЗА ────────────────────────────────────

STATUS_MESSAGES = {
    "preparing": "👨‍🍳 Ваш заказ #{} готовится!",
    "on_way":    "🚚 Ваш заказ #{} в пути! Ожидайте.",
    "ready":     "📦 Ваш заказ #{} готов к выдаче!",
    "delivered": "✅ Ваш заказ #{} доставлен! Спасибо.",
    "given":     "✅ Ваш заказ #{} выдан! Спасибо.",
}

STATUS_LABELS = {
    "preparing": "👨‍🍳 ГОТОВИТСЯ",
    "on_way":    "🚚 В ПУТИ",
    "ready":     "📦 ГОТОВ К ВЫДАЧЕ",
    "delivered": "✅ ДОСТАВЛЕН",
    "given":     "✅ ВЫДАН",
}


@router.callback_query(F.data.startswith("status_"))
async def update_status(callback: CallbackQuery, bot):
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return

    _, rest = callback.data.split("_", 1)
    status, order_id = rest.rsplit("_", 1)
    order_id = int(order_id)

    async with session_maker() as session:
        order = await session.get(Order, order_id)
        order.status = status
        await session.commit()
        client = await session.get(Client, order.client_id)
        telegram_id = client.telegram_id
        delivery_type = order.delivery_type
        msg_id = order.status_message_id

    status_text = {
        "preparing": "👨‍🍳 Заказ готовится...",
        "on_way":    "🚚 Заказ в пути! Ожидайте.",
        "ready":     "📦 Заказ готов к выдаче!",
        "delivered": "✅ Заказ доставлен! Спасибо.",
        "given":     "✅ Заказ выдан! Спасибо.",
    }

    final_statuses = ["delivered", "given"]
    text = f"{status_text[status]}"

    if status in final_statuses:
        # Редактируем последнее сообщение
        if msg_id:
            try:
                await bot.edit_message_text(
                    chat_id=telegram_id,
                    message_id=msg_id,
                    text=text
                )
            except Exception:
                await bot.send_message(telegram_id, text)

        # Отправляем оценку отдельным сообщением
        from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        rating_kb = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="😞", callback_data=f"rate_1_{order_id}"),
                InlineKeyboardButton(text="😐", callback_data=f"rate_2_{order_id}"),
                InlineKeyboardButton(text="🙂", callback_data=f"rate_3_{order_id}"),
                InlineKeyboardButton(text="😊", callback_data=f"rate_4_{order_id}"),
                InlineKeyboardButton(text="🤩", callback_data=f"rate_5_{order_id}"),
            ],
            [InlineKeyboardButton(text="Пропустить", callback_data=f"rate_0_{order_id}")]
        ])
        await bot.send_message(telegram_id, "Как вам доставка?", reply_markup=rating_kb)
        await callback.message.delete()

    else:
        # Редактируем существующее сообщение
        if msg_id:
            try:
                await bot.edit_message_text(
                    chat_id=telegram_id,
                    message_id=msg_id,
                    text=text
                )
            except Exception:
                pass

        await callback.message.edit_reply_markup(
            reply_markup=order_status_keyboard(order_id, delivery_type, status)
        )

# ─── REPLY КНОПКИ АДМИНА ───────────────────────────────

@router.message(F.text == "📦 Новые заказы")
async def new_orders_reply(message: Message):
    if not is_admin(message.from_user.id):
        return
    async with session_maker() as session:
        orders = await session.execute(
            select(Order).where(Order.status == "new")
        )
        orders = orders.scalars().all()

        if not orders:
            _track_msg(message.chat.id, await message.answer("📭 Новых заказов нет."))
            return

        for order in orders:
            client = await session.get(Client, order.client_id)
            items = await session.execute(
                select(OrderItem).where(OrderItem.order_id == order.id)
            )
            items = items.scalars().all()

            type_text = "🚚 Доставка" if order.delivery_type == "delivery" else "🏪 Самовывоз"
            address_line = f"\n📍 {order.delivery_address}" if order.delivery_type == "delivery" and order.delivery_address else ""
            requisites = _requisites_text(client)
            text = (
                f"📦 ЗАКАЗ #{order.id}\n"
                f"👤 {client.contact_name}\n"
                f"🏢 {client.organization}\n"
                f"📞 {client.phone}\n"
                f"{requisites}"
                f"━━━━━━━━━━━━━━━\n"
            )
            for item in items:
                text += f"• {item.product_name} — {item.quantity} шт\n"
            text += (
                f"━━━━━━━━━━━━━━━\n"
                f"{type_text}{address_line}\n"
                f"📅 {order.delivery_date} в {order.delivery_time}"
            )
            from keyboards import order_action
            _track_msg(message.chat.id, await message.answer(text, reply_markup=order_action(order.id)))


@router.message(F.text == "⚙️ Активные заказы")
async def active_orders_reply(message: Message):
    if not is_admin(message.from_user.id):
        return

    active = ["accepted", "preparing", "on_way", "ready"]

    async with session_maker() as session:
        orders = await session.execute(
            select(Order).where(Order.status.in_(active))
        )
        orders = orders.scalars().all()

        if not orders:
            _track_msg(message.chat.id, await message.answer("📭 Активных заказов нет."))
            return

        for order in orders:
            client = await session.get(Client, order.client_id)
            items = await session.execute(
                select(OrderItem).where(OrderItem.order_id == order.id)
            )
            items = items.scalars().all()

            type_text = "🚚 Доставка" if order.delivery_type == "delivery" else "🏪 Самовывоз"
            address_line = f"\n📍 {order.delivery_address}" if order.delivery_type == "delivery" and order.delivery_address else ""
            status_map = {
                "accepted":  "✅ Принят",
                "preparing": "👨‍🍳 Готовится",
                "on_way":    "🚚 В пути",
                "ready":     "📦 Готов к выдаче",
            }
            requisites = _requisites_text(client)
            text = (
                f"📦 Заказ #{order.id} — {status_map.get(order.status)}\n"
                f"👤 {client.contact_name}\n"
                f"🏢 {client.organization}\n"
                f"{requisites}"
                f"━━━━━━━━━━━━━━━\n"
            )
            for item in items:
                text += f"• {item.product_name} — {item.quantity} шт\n"
            text += (
                f"━━━━━━━━━━━━━━━\n"
                f"{type_text}{address_line}\n"
                f"📅 {order.delivery_date} в {order.delivery_time}"
            )
            _track_msg(message.chat.id, await message.answer(
                text,
                reply_markup=order_status_keyboard(order.id, order.delivery_type, order.status)
            ))


@router.message(F.text == "📋 История заказов")
async def all_orders_reply(message: Message):
    if not is_admin(message.from_user.id):
        return

    async with session_maker() as session:
        orders = await session.execute(select(Order))
        orders = orders.scalars().all()

        if not orders:
            _track_msg(message.chat.id, await message.answer("📭 Заказов пока нет."))
            return

        status_map = {
            "new":       "🟡 Новый",
            "accepted":  "✅ Принят",
            "rejected":  "❌ Отклонён",
            "cancelled": "🚫 Отменён",
            "preparing": "👨‍🍳 Готовится",
            "on_way":    "🚚 В пути",
            "ready":     "📦 Готов к выдаче",
            "delivered": "✅ Доставлен",
            "given":     "✅ Выдан",
        }

        chunks = ["📋 ВСЕ ЗАКАЗЫ:\n\n"]
        for order in orders:
            client = await session.get(Client, order.client_id)
            type_icon = "🚚" if order.delivery_type == "delivery" else "🏪"
            line = (
                f"{status_map.get(order.status)} #{order.id} — "
                f"{client.contact_name} — "
                f"{type_icon} {order.delivery_date}\n"
            )
            if len(chunks[-1]) + len(line) > 3500:
                chunks.append("")
            chunks[-1] += line

    for chunk in chunks:
        if chunk.strip():
            _track_msg(message.chat.id, await message.answer(chunk))


@router.message(F.text == "🍞 Управление меню")
async def manage_menu_reply(message: Message):
    if not is_admin(message.from_user.id):
        return
    _track_msg(message.chat.id, await message.answer("🍞 Управление меню:", reply_markup=manage_menu()))


# ─── ПОИСК ЗАКАЗА ──────────────────────────────────────

class SearchState(StatesGroup):
    waiting_order_id = State()


@router.message(F.text == "🔍 Найти заказ")
async def search_order_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.set_state(SearchState.waiting_order_id)
    _track_msg(message.chat.id, await message.answer("Введите номер заказа:"))


@router.message(SearchState.waiting_order_id)
async def search_order_result(message: Message, state: FSMContext):
    try:
        order_id = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Введите число — номер заказа.")
        return  # state не сбрасываем — пользователь может попробовать снова

    await state.clear()

    async with session_maker() as session:
        order = await session.get(Order, order_id)

        if not order:
            await message.answer(f"❌ Заказ #{order_id} не найден.")
            return

        client = await session.get(Client, order.client_id)
        items = await session.execute(
            select(OrderItem).where(OrderItem.order_id == order.id)
        )
        items = items.scalars().all()

    status_map = {
        "new":       "🟡 Новый",
        "accepted":  "✅ Принят",
        "rejected":  "❌ Отклонён",
        "cancelled": "🚫 Отменён",
        "preparing": "👨‍🍳 Готовится",
        "on_way":    "🚚 В пути",
        "ready":     "📦 Готов к выдаче",
        "delivered": "✅ Доставлен",
        "given":     "✅ Выдан",
    }

    type_text = "🚚 Доставка" if order.delivery_type == "delivery" else "🏪 Самовывоз"
    address_line = f"\n📍 {order.delivery_address}" if order.delivery_type == "delivery" and order.delivery_address else ""
    requisites = _requisites_text(client)
    text = (
        f"🔍 ЗАКАЗ #{order.id}\n"
        f"👤 {client.contact_name}\n"
        f"🏢 {client.organization}\n"
        f"📞 {client.phone}\n"
        f"{requisites}"
        f"📊 {status_map.get(order.status, order.status)}\n"
        f"━━━━━━━━━━━━━━━\n"
    )
    for item in items:
        text += f"• {item.product_name} — {item.quantity} шт\n"
    text += (
        f"━━━━━━━━━━━━━━━\n"
        f"{type_text}{address_line}\n"
        f"📅 {order.delivery_date} в {order.delivery_time}\n"
        f"🕐 Создан: {order.created_at.strftime('%d.%m.%Y %H:%M')}"
    )

    _track_msg(message.chat.id, await message.answer(text))

# ─── ФИЛЬТР ПО ДАТЕ ────────────────────────────────────

class FilterState(StatesGroup):
    waiting_date = State()


@router.message(F.text == "📅 Заказы по дате")
async def filter_by_date_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.set_state(FilterState.waiting_date)
    _track_msg(message.chat.id, await message.answer("Введите дату в формате ДД.ММ.ГГГГ\nНапример: 06.03.2026"))


@router.message(FilterState.waiting_date)
async def filter_by_date_result(message: Message, state: FSMContext):
    date_input = message.text.strip()

    # Валидация формата
    try:
        datetime.strptime(date_input, "%d.%m.%Y")
    except ValueError:
        await message.answer("❌ Неверный формат. Введите дату как 06.03.2026")
        return  # state не сбрасываем — пользователь может попробовать снова

    await state.clear()

    async with session_maker() as session:
        orders = await session.execute(
            select(Order).where(Order.delivery_date == date_input)
        )
        orders = orders.scalars().all()

    if not orders:
        _track_msg(message.chat.id, await message.answer(f"📭 Заказов на {date_input} нет."))
        return

    status_map = {
        "new":       "🟡",
        "accepted":  "✅",
        "rejected":  "❌",
        "cancelled": "🚫",
        "preparing": "👨‍🍳",
        "on_way":    "🚚",
        "ready":     "📦",
        "delivered": "☑️",
        "given":     "☑️",
    }

    text = f"📅 ЗАКАЗЫ НА {date_input}:\n\n"
    async with session_maker() as session:
        for order in orders:
            client = await session.get(Client, order.client_id)
            type_icon = "🚚" if order.delivery_type == "delivery" else "🏪"
            text += (
                f"{status_map.get(order.status, '❓')} #{order.id} — "
                f"{client.contact_name} — "
                f"{type_icon} {order.delivery_time}\n"
            )

    _track_msg(message.chat.id, await message.answer(text))

# ─── СТАТИСТИКА ────────────────────────────────────────

from sqlalchemy import func


@router.message(F.text == "📊 Статистика")
async def statistics(message: Message):
    if not is_admin(message.from_user.id):
        return

    today = datetime.now().strftime("%d.%m.%Y")

    # Начало недели (понедельник)
    now = datetime.now()
    week_start = (now - timedelta(days=now.weekday())).strftime("%d.%m.%Y")

    async with session_maker() as session:
        # Всего заказов
        total = await session.execute(select(func.count(Order.id)))
        total = total.scalar()

        # За сегодня
        today_orders = await session.execute(
            select(func.count(Order.id)).where(Order.delivery_date == today)
        )
        today_count = today_orders.scalar()

        # Активные прямо сейчас
        active = await session.execute(
            select(func.count(Order.id)).where(
                Order.status.in_(["accepted", "preparing", "on_way", "ready"])
            )
        )
        active_count = active.scalar()

        # Отменённых/отклонённых
        cancelled = await session.execute(
            select(func.count(Order.id)).where(
                Order.status.in_(["rejected", "cancelled"])
            )
        )
        cancelled_count = cancelled.scalar()

        # Завершённых
        done = await session.execute(
            select(func.count(Order.id)).where(
                Order.status.in_(["delivered", "given"])
            )
        )
        done_count = done.scalar()

        # За текущую неделю (по дате доставки — парсим строки перед сравнением)
        all_orders_list = await session.execute(select(Order))
        week_start_dt = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        week_count = sum(
            1 for o in all_orders_list.scalars().all()
            if datetime.strptime(o.delivery_date, "%d.%m.%Y") >= week_start_dt
        )

    text = (
        f"📊 СТАТИСТИКА\n"
        f"━━━━━━━━━━━━━━━\n"
        f"📦 Всего заказов: {total}\n"
        f"📅 Сегодня ({today}): {today_count}\n"
        f"📆 Эта неделя (с {week_start}): {week_count}\n"
        f"━━━━━━━━━━━━━━━\n"
        f"⚙️ Активных: {active_count}\n"
        f"☑️ Выполнено: {done_count}\n"
        f"❌ Отменено/отклонено: {cancelled_count}\n"
    )

    _track_msg(message.chat.id, await message.answer(text))

# ─── ЭКСПОРТ В EXCEL ───────────────────────────────────

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


@router.message(F.text == "📥 Экспорт в Excel")
async def export_excel(message: Message):
    if not is_admin(message.from_user.id):
        return

    async with session_maker() as session:
        orders = await session.execute(select(Order))
        orders = orders.scalars().all()

        if not orders:
            await message.answer("📭 Заказов нет.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Заказы"

        # Заголовки
        headers = ["№", "Клиент", "Организация", "Телефон",
                   "ИНН", "Р/С", "МФО", "Банк", "Юр. адрес",
                   "Продукты", "Тип", "Адрес", "Дата", "Время", "Статус", "Комментарий"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFC300")
            cell.alignment = Alignment(horizontal="center")

        status_map = {
            "new": "Новый", "accepted": "Принят", "rejected": "Отклонён",
            "cancelled": "Отменён", "preparing": "Готовится",
            "on_way": "В пути", "ready": "Готов", "delivered": "Доставлен", "given": "Выдан"
        }

        for row, order in enumerate(orders, 2):
            client = await session.get(Client, order.client_id)
            items = await session.execute(
                select(OrderItem).where(OrderItem.order_id == order.id)
            )
            items = items.scalars().all()
            products_text = ", ".join([f"{i.product_name} x{i.quantity}" for i in items])
            type_text = "Доставка" if order.delivery_type == "delivery" else "Самовывоз"

            ws.append([
                order.id,
                client.contact_name,
                client.organization,
                client.phone,
                client.inn or "",
                client.bank_account or "",
                client.mfo or "",
                client.bank_name or "",
                client.legal_address or "",
                products_text,
                type_text,
                order.delivery_address or "",
                order.delivery_date,
                order.delivery_time,
                status_map.get(order.status, order.status),
                order.comment or ""
            ])

        # Ширина колонок
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 22
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 30
        ws.column_dimensions["J"].width = 35
        ws.column_dimensions["K"].width = 12
        ws.column_dimensions["L"].width = 30
        ws.column_dimensions["M"].width = 12
        ws.column_dimensions["N"].width = 8
        ws.column_dimensions["O"].width = 15
        ws.column_dimensions["P"].width = 30

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"../orders_{timestamp}.xlsx")
    path = os.path.normpath(path)
    wb.save(path)

    from aiogram.types import FSInputFile
    file = FSInputFile(path)
    await message.answer_document(file, caption="📊 Все заказы")
    os.remove(path)


# ─── РАССЫЛКА ──────────────────────────────────────────

class BroadcastState(StatesGroup):
    waiting_message = State()


@router.message(F.text == "📢 Рассылка")
async def broadcast_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.set_state(BroadcastState.waiting_message)
    _track_msg(message.chat.id, await message.answer(
        "📢 Введите текст рассылки.\n"
        "Получат все зарегистрированные клиенты.\n\n"
        "Или напишите /cancel для отмены:"
    ))


@router.message(F.text == "/cancel", BroadcastState.waiting_message)
async def broadcast_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("❌ Рассылка отменена.")


@router.message(BroadcastState.waiting_message)
async def broadcast_send(message: Message, state: FSMContext, bot):
    await state.clear()
    text = message.text

    async with session_maker() as session:
        clients = await session.execute(select(Client))
        clients = clients.scalars().all()

    success = 0
    failed = 0

    for i, client in enumerate(clients):
        try:
            await bot.send_message(
                client.telegram_id,
                f"📢 Сообщение от пекарни:\n\n{text}"
            )
            success += 1
        except Exception:
            failed += 1
        # Telegram лимит: 30 сообщений/сек — держим ~25/сек с запасом
        if i % 25 == 24:
            await asyncio.sleep(1)

    await message.answer(f"✅ Рассылка отправлена {success} клиентам!")


# ─── ДОГОВОРНЫЕ ЦЕНЫ ───────────────────────────────────

class PriceState(StatesGroup):
    selecting_client = State()
    selecting_product = State()
    entering_price = State()


@router.message(F.text == "💰 Договорные цены")
async def custom_prices_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    async with session_maker() as session:
        clients = await session.execute(select(Client))
        clients = clients.scalars().all()

    if not clients:
        await message.answer("😔 Клиентов пока нет.")
        return

    text = "👤 Выберите клиента (введите номер):\n\n"
    for i, c in enumerate(clients, 1):
        text += f"{i}. {c.contact_name} — {c.organization}\n"

    await state.update_data(clients=[(c.id, c.contact_name, c.organization) for c in clients])
    await state.set_state(PriceState.selecting_client)
    _track_msg(message.chat.id, await message.answer(text))


@router.message(PriceState.selecting_client)
async def select_client_for_price(message: Message, state: FSMContext):
    data = await state.get_data()
    clients = data.get("clients")
    if not clients:
        await state.clear()
        return

    try:
        index = int(message.text) - 1
        if index < 0 or index >= len(clients):
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите правильный номер")
        return

    client_id, contact_name, organization = clients[index]
    await state.update_data(selected_client_id=client_id, selected_client_name=contact_name)

    async with session_maker() as session:
        products = await session.execute(select(Product))
        products = products.scalars().all()

        # Текущие договорные цены клиента
        custom = await session.execute(
            select(ClientPrice).where(ClientPrice.client_id == client_id)
        )
        custom = {cp.product_id: cp.price for cp in custom.scalars().all()}

    text = f"🍞 Продукты для {contact_name}:\n\n"
    for i, p in enumerate(products, 1):
        if p.id in custom:
            text += f"{i}. {p.name} — {p.price} сум (договорная: {custom[p.id]} сум)\n"
        else:
            text += f"{i}. {p.name} — {p.price} сум\n"
    text += "\nВведите номер продукта чтобы изменить цену:"

    await state.update_data(products=[(p.id, p.name, p.price) for p in products])
    await state.set_state(PriceState.selecting_product)
    await message.answer(text)


@router.message(PriceState.selecting_product)
async def select_product_for_price(message: Message, state: FSMContext):
    data = await state.get_data()
    products = data["products"]

    try:
        index = int(message.text) - 1
        if index < 0 or index >= len(products):
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите правильный номер")
        return

    product_id, product_name, standard_price = products[index]
    await state.update_data(selected_product_id=product_id, selected_product_name=product_name)
    await state.set_state(PriceState.entering_price)
    await message.answer(
        f"💰 Стандартная цена {product_name}: {standard_price} сум\n"
        f"Введите договорную цену или 0 чтобы сбросить до стандартной:"
    )


@router.message(PriceState.entering_price)
async def save_custom_price(message: Message, state: FSMContext):
    try:
        price = int(message.text)
        if price < 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите правильную цену числом")
        return

    data = await state.get_data()
    client_id = data["selected_client_id"]
    client_name = data["selected_client_name"]
    product_id = data["selected_product_id"]
    product_name = data["selected_product_name"]

    async with session_maker() as session:
        existing = await session.execute(
            select(ClientPrice).where(
                ClientPrice.client_id == client_id,
                ClientPrice.product_id == product_id
            )
        )
        existing = existing.scalar_one_or_none()

        if price == 0:
            # Сбрасываем до стандартной
            if existing:
                await session.delete(existing)
            await session.commit()
            await state.clear()
            await message.answer(f"✅ Цена {product_name} для {client_name} сброшена до стандартной.")
        else:
            if existing:
                existing.price = price
            else:
                session.add(ClientPrice(
                    client_id=client_id,
                    product_id=product_id,
                    price=price
                ))
            await session.commit()
            await state.clear()
            await message.answer(f"✅ Договорная цена {product_name} для {client_name}: {price} сум")

# ─── СПИСОК КЛИЕНТОВ ───────────────────────────────────

@router.message(F.text == "👥 Клиенты")
async def clients_list(message: Message):
    if not is_admin(message.from_user.id):
        return

    async with session_maker() as session:
        clients = await session.execute(select(Client))
        clients = clients.scalars().all()

    if not clients:
        _track_msg(message.chat.id, await message.answer("Клиентов пока нет."))
        return

    chunks = [f"<b>КЛИЕНТЫ · {len(clients)}</b>\n━━━━━━━━━━━━━━━━━━━━\n"]
    for i, c in enumerate(clients, 1):
        card = f"<b>{i}. {c.contact_name}</b>\n"
        card += f"🏢 {c.organization}\n"
        card += f"📞 {c.phone}\n"
        if c.inn:
            card += f"🏦 ИНН: {c.inn}\n"
        if c.bank_account:
            mfo = f" | МФО: {c.mfo}" if c.mfo else ""
            card += f"💳 Р/С: {c.bank_account}{mfo}\n"
        if c.bank_name:
            card += f"🏛 {c.bank_name}\n"
        if c.legal_address:
            card += f"📍 {c.legal_address}\n"
        card += "━━━━━━━━━━━━━━━━━━━━\n"

        if len(chunks[-1]) + len(card) > 3800:
            chunks.append("")
        chunks[-1] += card

    for chunk in chunks:
        _track_msg(message.chat.id, await message.answer(chunk, parse_mode="HTML"))


@router.message(F.text == "⏸ Пауза")
async def pause_bot(message: Message):
    if not is_admin(message.from_user.id):
        return
    config.bot_paused = True
    _track_msg(message.chat.id, await message.answer(
        "⏸ Бот на паузе. Клиенты не могут делать заказы.\n"
        "Нажми '▶️ Возобновить' чтобы включить обратно."
    ))


@router.message(F.text == "▶️ Возобновить")
async def resume_bot(message: Message):
    if not is_admin(message.from_user.id):
        return
    config.bot_paused = False
    _track_msg(message.chat.id, await message.answer("▶️ Бот возобновлён. Клиенты снова могут делать заказы."))